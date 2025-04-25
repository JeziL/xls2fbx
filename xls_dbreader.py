import re
from math import modf
from openpyxl import load_workbook


class Signal:
    """单个信号字段类"""

    def __init__(
        self,
        name,
        desc,
        data_type,
        start_bit,
        length,
        factor=None,
        offset=None,
        text_table=None,
    ) -> None:
        self.name = name
        self.desc = desc
        self.data_type = data_type
        self.start_bit = start_bit
        self.length = length
        self.factor = factor
        self.offset = offset
        self.text_table = text_table
        self.fbx_id = ""
        self.coding_fbx_id = ""

    def __str__(self) -> str:
        return f"<Signal {self.name}>"

    def __repr__(self) -> str:
        return self.__str__()


class Frame:
    """FlexRay 帧类"""

    def __init__(
        self,
        name="",
        desc="",
        length=32,
        frame_type="static",
        slot=1,
        base_cycle=0,
        cycle_rep=1,
    ) -> None:
        self.name = name
        self.desc = desc
        self.length = length
        self.frame_type = frame_type
        self.slot = slot
        self.base_cycle = base_cycle
        self.cycle_rep = cycle_rep
        self.signals = []
        self.is_startup = False
        self.fbx_id = ""
        self.ft_fbx_id = ""
        self.channels = [True, True]

    def read_signals(self, worksheet):
        """解析帧内包含的所有字段"""
        for row in worksheet.iter_rows(min_row=5, max_col=8, max_row=100):
            if not row[0].value:
                break
            bit_offset, start_byte = modf(float(row[3].value))
            start_bit = int(start_byte) * 8 + round(bit_offset * 10)
            if int(row[4].value) > 8:
                start_bit = (start_bit + int(row[4].value) - 1) // 8 * 8
            signal = Signal(
                row[0].value.strip(),
                row[1].value.strip(),
                row[2].value.strip(),
                start_bit,
                row[4].value,
            )
            if row[5].value or row[6].value:
                signal.factor = row[5].value if row[5].value else 1
                signal.offset = row[6].value if row[6].value else 0
            if row[7].value:
                text_table = {}
                for tt in row[7].value.splitlines():
                    val = int(re.split("[:：]", tt)[0].strip(), 0)
                    desc = re.split("[:：]", tt)[1].strip()
                    text_table[val] = desc
                signal.text_table = text_table
            self.signals.append(signal)

    def __str__(self) -> str:
        return f"<Frame {self.name}: {self.desc}>"

    def __repr__(self) -> str:
        return self.__str__()


class ECU:
    """ECU 节点类"""

    def __init__(self, name, desc, channel_A, channel_B) -> None:
        self.name = name
        self.desc = desc
        self.channels = [channel_A, channel_B]
        self.frames = []
        self.fbx_id = ""
        self.ctrl_fbx_id = ""

    def __str__(self) -> str:
        return f"<ECU {self.name}: {self.desc}>"

    def __repr__(self) -> str:
        return self.__str__()


class XLSDatabase:
    """FlexRay 总线数据库类"""

    def __init__(self, filename) -> None:
        self.wb = load_workbook(filename)
        self.ecus = []
        self.read_ECUs()
        for ws in self.wb.worksheets:
            if ws.title == "节点列表":
                continue
            self.read_Frame(ws)

    def read_ECUs(self) -> None:
        """读取所有 ECU 节点"""
        ws = self.wb["节点列表"]
        for row in ws.iter_rows(min_row=2, max_col=3, max_row=50):
            if not row[0].value:
                break
            self.ecus.append(
                ECU(
                    row[0].value, row[1].value, "A" in row[2].value, "B" in row[2].value
                )
            )

    def read_Frame(self, worksheet) -> None:
        """读取所有帧，并与 ECU 节点绑定"""
        target_ecu = None
        for ecu in self.ecus:
            if ecu.name == worksheet["H1"].value:
                target_ecu = ecu
        if not target_ecu:
            target_ecu = ECU(worksheet["H1"].value, "", True, True)
            self.ecus.append(target_ecu)

        frame = Frame()
        frame.name = worksheet["B1"].value
        frame.desc = worksheet["D1"].value
        frame.length = worksheet["F1"].value
        frame.frame_type = "static" if worksheet["B2"].value == "静态帧" else "dynamic"
        frame.slot = worksheet["D2"].value
        frame.base_cycle = worksheet["F2"].value
        frame.cycle_rep = worksheet["H2"].value
        frame.is_startup = worksheet["J1"].value.strip() == "是"
        frame.channels = ["A" in worksheet["J2"].value, "B" in worksheet["J2"].value]

        frame.read_signals(worksheet)
        target_ecu.frames.append(frame)
